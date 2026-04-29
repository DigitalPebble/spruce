#!/usr/bin/env python3
"""
Downloads Aqueduct 4.0 country/province baseline water stress rankings and maps
them to cloud regions defined in cloud_regions.json.

Resolution order per cloud region (indicator: bws, weight: Tot):
  1. province_baseline — matched via Nominatim reverse geocode (state/province name)
  2. country_baseline  — matched by country name

Output columns: provider,region,water_stress_cat
  cat: 0=Low, 1=Low-Medium, 2=Medium-High, 3=High, 4=Extremely High, -9999=NoData

Usage:
  ./fetch_aqueduct_water_stress.py [cloud_regions.json] [output.csv]

Requires: Python 3.8+, openpyxl  (pip install openpyxl)
"""

import csv
import json
import os
import sys
import time
import unicodedata
import urllib.error
import urllib.request
import zipfile
from io import BytesIO

try:
    import openpyxl
except ImportError:
    sys.exit("error: openpyxl not installed — run: pip install openpyxl")

AQUEDUCT_URL = "https://files.wri.org/aqueduct/aqueduct-4-0-country-rankings.zip"
XLSX_NAME = "Aqueduct40_rankings_download_Y2023M07D05/Aqueduct40_rankings_download_Y2023M07D05.xlsx"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/reverse"
NOMINATIM_UA = "aqueduct-cloud-region-script/1.0 (contact: spruce@digitalpebble.com)"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

# Country name differences between cloud_regions.json and Aqueduct (GADM)
COUNTRY_NAME_MAP = {
    "Taiwan": "Taiwan",        # Aqueduct has no TWN row — will fall back gracefully
    "Hong Kong": "Hong Kong",
}

# Aqueduct uses these country names; cloud_regions.json uses the keys
CLOUD_TO_AQUEDUCT_COUNTRY = {
    "South Korea": "South Korea",   # matches as-is
    "United Kingdom": "United Kingdom",
    "United States": "United States",
    "United Arab Emirates": "United Arab Emirates",
}


def normalize(s: str) -> str:
    """Lowercase, strip accents, collapse whitespace for fuzzy name matching."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(ascii_s.lower().split())


def load_aqueduct(xlsx_path: str):
    """
    Returns two dicts:
      country_cat:  name_0 (normalised) → cat
      province_cat: (name_0_normalised, name_1_normalised) → cat
    Only includes rows where indicator_name='bws' and weight='Tot'.
    Rows with cat=-9999 (NoData) are kept so we can distinguish from 'not found'.
    """
    print(f"Parsing {xlsx_path} …", file=sys.stderr)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    country_cat: dict[str, int] = {}
    ws_c = wb["country_baseline"]
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        # gid_0, name_0, indicator_name, weight, score, score_ranked, cat, label, …
        if row[2] == "bws" and row[3] == "Tot" and row[6] is not None:
            country_cat[normalize(row[1])] = int(row[6])

    province_cat: dict[tuple[str, str], int] = {}
    ws_p = wb["province_baseline"]
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        # gid_0, gid_1, name_0, name_1, indicator_name, weight, score, score_ranked, cat, …
        if row[4] == "bws" and row[5] == "Tot" and row[8] is not None:
            key = (normalize(row[2]), normalize(row[3]))
            province_cat[key] = int(row[8])

    wb.close()
    print(
        f"  {len(country_cat)} countries, {len(province_cat)} provinces loaded",
        file=sys.stderr,
    )
    return country_cat, province_cat


def download_xlsx(cache_dir: str) -> str:
    """Download and extract the Aqueduct xlsx, returning its local path."""
    xlsx_path = os.path.join(cache_dir, os.path.basename(XLSX_NAME))
    if os.path.exists(xlsx_path):
        print(f"Using cached {xlsx_path}", file=sys.stderr)
        return xlsx_path

    print(f"Downloading {AQUEDUCT_URL} …", file=sys.stderr)
    req = urllib.request.Request(AQUEDUCT_URL, headers={"User-Agent": NOMINATIM_UA})
    with urllib.request.urlopen(req) as resp:
        data = resp.read()

    with zipfile.ZipFile(BytesIO(data)) as zf:
        zf.extract(XLSX_NAME, path=cache_dir)

    # Move extracted file to cache root for easy access
    extracted = os.path.join(cache_dir, XLSX_NAME)
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    # extracted is already at a deep path; just use it directly
    print(f"Extracted to {extracted}", file=sys.stderr)
    return extracted


def _is_ascii(s: str) -> bool:
    try:
        s.encode("ascii")
        return True
    except UnicodeEncodeError:
        return False


def load_geocache(path: str) -> dict[tuple[str, str], str]:
    """Load TSV geocache: lat<TAB>lon<TAB>state_name (empty string = no result).

    Entries with non-ASCII state names are discarded so they are re-fetched
    with the corrected Accept-Language: en header.
    """
    cache: dict[tuple[str, str], str] = {}
    if not os.path.exists(path):
        return cache
    with open(path) as f:
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) >= 3:
                state = parts[2]
                if state and not _is_ascii(state):
                    continue  # stale entry fetched without Accept-Language: en
                cache[(parts[0], parts[1])] = state
    return cache


def save_geocache_entry(path: str, lat: str, lon: str, state: str) -> None:
    with open(path, "a") as f:
        f.write(f"{lat}\t{lon}\t{state}\n")


# Trailing administrative suffixes returned by Nominatim (English) but absent in Aqueduct
_ADMIN_SUFFIXES = (
    " prefecture", " province", " county", " region",
    " governorate", " district", " municipality",
    " metropolis", " autonomous region",
)


def _strip_admin_suffix(name: str) -> str:
    """Remove common administrative suffixes so 'Kanagawa Prefecture' → 'Kanagawa'."""
    lower = name.lower()
    for suffix in _ADMIN_SUFFIXES:
        if lower.endswith(suffix):
            return name[: len(name) - len(suffix)].strip()
    return name


def reverse_geocode_state(lat: str, lon: str) -> str:
    """
    Call Nominatim to get the state/province name for (lat, lon).
    Returns empty string if nothing useful is found.
    """
    url = f"{NOMINATIM_URL}?format=jsonv2&zoom=6&lat={lat}&lon={lon}"
    req = urllib.request.Request(
        url,
        headers={"User-Agent": NOMINATIM_UA, "Accept-Language": "en"},
    )
    try:
        time.sleep(1.1)  # Nominatim: max 1 req/sec
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
    except (urllib.error.URLError, json.JSONDecodeError) as exc:
        print(f"  geocode warning: {exc}", file=sys.stderr)
        return ""

    addr = data.get("address", {})
    # Try fields in priority order; different countries use different keys
    for field in ("state", "province", "region", "county"):
        val = addr.get(field, "")
        if val:
            return _strip_admin_suffix(val)
    return ""


def lookup_water_stress(
    country: str,
    lat: str,
    lon: str,
    country_cat: dict,
    province_cat: dict,
    geocache: dict,
    geocache_path: str,
) -> int | None:
    """
    Returns water stress cat (int) or None if not found.
    Tries province first, then country.
    """
    norm_country = normalize(country)

    # --- Province resolution ---
    key = (lat, lon)
    if key in geocache:
        state = geocache[key]
    else:
        state = reverse_geocode_state(lat, lon)
        geocache[key] = state
        save_geocache_entry(geocache_path, lat, lon, state)

    if state:
        pkey = (norm_country, normalize(_strip_admin_suffix(state)))
        if pkey in province_cat:
            cat = province_cat[pkey]
            return cat if cat != -9999 else None

    # --- Country fallback ---
    if norm_country in country_cat:
        cat = country_cat[norm_country]
        return cat if cat != -9999 else None

    return None


def main():
    cloud_regions_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        PROJECT_ROOT, "src", "main", "resources", "cloud_regions.json"
    )
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        PROJECT_ROOT, "src", "main", "resources", "water-stress.csv"
    )
    geocache_path = os.environ.get(
        "GEOCACHE", os.path.join(os.path.dirname(output_path), "geocode_cache.tsv")
    )
    cache_dir = os.path.join(SCRIPT_DIR, ".aqueduct_cache")
    os.makedirs(cache_dir, exist_ok=True)

    if not os.path.exists(cloud_regions_path):
        sys.exit(f"error: {cloud_regions_path} not found")

    xlsx_path = download_xlsx(cache_dir)
    country_cat, province_cat = load_aqueduct(xlsx_path)

    with open(cloud_regions_path) as f:
        cloud_regions = json.load(f)

    geocache = load_geocache(geocache_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    rows: list[tuple[str, str, str]] = []
    providers = ["aws", "gcp", "azure"]

    for provider in providers:
        regions = cloud_regions.get(provider, {}).get("cloud_regions", {})
        for region_code, info in regions.items():
            country = info.get("country", "")
            lat = info.get("latitude", "")
            lon = info.get("longitude", "")
            if not country or not lat or not lon:
                continue

            cat = lookup_water_stress(
                country, lat, lon, country_cat, province_cat, geocache, geocache_path
            )
            if cat is not None:
                rows.append((provider, region_code, str(cat)))
                print(f"  {provider}:{region_code} ({country}) → {cat}", file=sys.stderr)
            else:
                print(
                    f"  {provider}:{region_code} ({country}) → no data", file=sys.stderr
                )

    with open(output_path, "w", newline="") as f:
        f.write("# Aqueduct 4.0 baseline water stress (bws, total weight)\n")
        f.write("# https://www.wri.org/data/aqueduct-global-maps-40-data\n")
        f.write("# cat: 0=Low, 1=Low-Medium, 2=Medium-High, 3=High, 4=Extremely High\n")
        writer = csv.writer(f)
        writer.writerow(["provider", "region", "water_stress_cat"])
        writer.writerows(sorted(rows))

    print(f"Wrote {output_path} ({len(rows)} rows)", file=sys.stderr)


if __name__ == "__main__":
    main()
